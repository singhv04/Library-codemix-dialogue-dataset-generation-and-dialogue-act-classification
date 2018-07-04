# -*- coding: utf-8 -*-
"""
Created on Tue Jun 26 19:47:30 2018

@author: hp
"""

# -*- coding: utf-8 -*-
"""
Created on Wed Jun 20 16:21:23 2018

@author: hp
"""



from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\dialogue act classiification\\q\\m\\kitni_m.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']
i=1
s='A'+str(i)
print(sheet_ranges[s].value)

#getting all books
s=""
m_kitni_q=[]
for i in range (1,47):
    s='A'+str(i)
    print(sheet_ranges[s].value)
    m_kitni_q.append(sheet_ranges[s].value)
    
from isc_tokenizer import Tokenizer
tk = Tokenizer(lang='hin')

ontologies=["मेंबर कोड","मेंबरकोड","कोड","memberCode","mcode","एम कोड",
            "name","नाम","मेंबर नेम","मेंबरनेम","मेंबर का नाम","membername",
            "thesis","book","थीसिस","ebook",
            "mail","email","contact","मेल","ईमेल","मेल id","ईमेल id","कांटेक्ट",
            "phone no","mobile number","नंबर","contact","फोन नंबर","मोबाइल नंबर","कांटेक्ट नंबर","कांटेक्ट"
            "groupcode","group","कोड","ग्रुप कोड","ग्रुपकोड","ग्रुप",
            "dues","due","ड्यू","बकाया","बाकी",
            "status","available","availability","मिल सकती","मिलेगी","स्टेटस","अवेलेबल","अवेलेबिलिटी",
            "accno","accesion no","accession number","अक्सेशन नंबर ",
            "title","name","नाम","टाइटल",
            "author","writer","लेखक","राइटर","ऑथर","लिखी","रचना","रची",
            "इशू डेट","इशू","ली","issue date","issuedt","issue","तारिख",
            "duedt","due date","due","ड्यू","ड्यू डेट","तारिख","submit","submission","सबमिट","सबमिशन",
            "maximum","max","मैक्स","मैक्सिमम","सबसे अधिक",
            "publication","पब्लिकेशन","publish","पब्लिश"]


stopwords=['कब','का','की','कि','के','को','है','हैं','था','या','सब','लिए','में','लाइब्रेरी','धन्यवाद','और','सबसे',
           'सारी','क्या','मिल','सकता','नाम','मिलेगा','आप','बता','सकते','सकती','जानकारी','मिलेगी','हो','बताइए',
           'दीजिए','किसी','ने','कराई','पर','सी','जो','जिनका','जिनके','जिनकी','जिनको','जिसका','जिसके','जिसकी',
           'जिसने','जिसको','जिन्हें','जिन्होंने','किसका','किसने','जिस','किस','किसके','जिसमें','जिसमे',
           'कितना','कितनी','कितने','कितनों','कोई']

pronouns_refering_previous=["इनका","उनका","इस","इसका","उन",'उसका','उस','उसके','वह','यह','इनका',
                            'इसका','उसका','इनके','उनके','इसके','उसमें','इनकी','उनकी','इसकी','उसकी',
                            'इस','उस','ये','वो','इन्होने','उन्होंने']
    
    
not_noun=['कब','का','की','कि','के','को','है','हैं','था','या','सब','लिए','में',
          'लाइब्रेरी','धन्यवाद','और','सबसे','सारी','क्या','मिल','सकता','नाम','मिलेगा',
          'आप','बता','सकते','सकती','जानकारी','मिलेगी','हो','बताइए','दीजिए','किसी',
          'ने','कराई','पर','सी','से','है']


z_n=[]
z_o=[]
z_p=[]

for a in range(0,len(m_kitni_q)):
    temp_str=m_kitni_q[a]
    pos=temp_str.find("कितनी")#question word postion
    #print(len("कितनी"))
    temp_str=temp_str[pos+len("कितनी"):]
    print(temp_str)
    test_str=tk.tokenize(temp_str)
    onto_found=[]
    for i in range(0,len(test_str)):
        for j in range(0,len(test_str)):
            if test_str[j]>='a' and test_str[j]<='z'or test_str[j]>='A' and test_str[j]<='Z':
                test_str[j]=test_str[j].lower()          
        if test_str[i] in ontologies:
            onto_found.append(test_str[i])
    z_o.append(onto_found)
    
for i in range(0,len(z_o)):
    if z_o[i] == []:
        z_o[i]=["DUES"]



#writing output to a file
import xlwt

# Initialize a workbook 
book = xlwt.Workbook(encoding="utf-8")

# Add a sheet to the workbook 
sheet1 = book.add_sheet("Python Sheet 1") 

row = 1
col = 0

for item in (m_kitni_q):
    sheet1.write(row, 0, item) 
    row += 1

row = 1
for item in (z_o):
    sheet1.write(row, 1, item) 
    row += 1
   
# Save the workbook 
book.save("info_asked_kitni_m.xls")