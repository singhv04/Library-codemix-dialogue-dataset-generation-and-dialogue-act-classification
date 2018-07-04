from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\bag_of_words_3.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']

#getting all bag_noun
s=""
bag_noun=[]
for i in range (2,11):
    s='B'+str(i)
    print(sheet_ranges[s].value)
    bag_noun.append(sheet_ranges[s].value)
    
    
#getting all bag_verb
s=""
bag_verb=[]
for i in range (2,8):
    s='C'+str(i)
    print(sheet_ranges[s].value)
    bag_verb.append(sheet_ranges[s].value)
    
#getting all bag_verb
s=""
bag_window=[]
for i in range (2,48):
    s='E'+str(i)
    print(sheet_ranges[s].value)
    bag_window.append(sheet_ranges[s].value)
   
   
#getting bag of words
bag_of_words=[]
wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\bag_of_words_1.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']
s=""
for i in range (2,156):
    s='A'+str(i)
    print(sheet_ranges[s].value)
    bag_of_words.append(sheet_ranges[s].value)
    
#bag_of_words_already tokenized with their index number
 
#we need to tokenize unigrams of the dialoguues in our train dataset corresponding to the bag_of_words tokens
#unigram 
token_unigrams=[]
unigram=[]
wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\feature_extracted.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']
s=""
temp_str=""
temp_list=[]
for i in range (2,48):
    s='C'+str(i)
    #print(sheet_ranges[s].value)
    temp_str=sheet_ranges[s].value
    temp_str=temp_str.strip()
    temp_str=temp_str.lower()
    print(temp_str)
    temp_list=temp_str.split(" ")
    unigram.append(temp_list)


#unigrams without stopwords    
from nltk.corpus import stopwords  
stop_words = set(stopwords.words('english'))  
unigrams_without_stopwords=[]
for i in range(0,len(unigram)):
    words=unigram[i]
    unigramsFiltered = []
    for w in words:
        if w not in stop_words:
            unigramsFiltered.append(w)
    unigrams_without_stopwords.append(unigramsFiltered)
    

    
#unigram_tokens  
from nltk.stem.porter import PorterStemmer
ps = PorterStemmer()
from nltk.corpus import stopwords
unigram_tokens=[]
pos=0
temp_tokens=[]#for one line
for i in range(0,len(unigram)):
    for j in range(0,len(unigram[i])):
        unigram[i][j] = ps.stem(unigram[i][j]) 
        print(unigram[i][j])
        if unigram[i][j] in bag_of_words:
            pos=bag_of_words.index(unigram[i][j])
            print(pos)
            temp_tokens.append(pos)
    unigram_tokens.append(temp_tokens)
    temp_tokens=[]
            
    
    


#verb in training sets
#AS THERE ARE ONLY ONE VERB IN EACH STATEMENT SO NO NEED TO PRE-PROCESS AND SPLIT THEM 
wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\feature_extracted.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']
s=""
verbs=[]
for i in range (2,48):
    s='E'+str(i)
    #print(sheet_ranges[s].value)
    temp_str=sheet_ranges[s].value
    print(temp_str)
    if temp_str is not None:
        verbs.append(temp_str)
    else:
        verbs.append("")
        
#verbs tokenize
from nltk.stem.porter import PorterStemmer
ps = PorterStemmer()
pos=0
temp=""
verbs_tokenize=[]
for i in range(0,len(verbs)):
    temp=ps.stem(verbs[i])
    #print(temp)
    temp=temp.strip()
    if temp in bag_verb:
        #print(verbs[i])
        pos=bag_verb.index(temp)
        verbs_tokenize.append(pos)
        print(pos)
    else:
        verbs_tokenize.append("")
        
   
    





#noun in training set
wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\feature_extracted.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']
s=""
nouns=[]
for i in range (2,48):
    s='D'+str(i)
    #print(sheet_ranges[s].value)
    temp_str=sheet_ranges[s].value
    print(temp_str)
    if temp_str is not None:
        nouns.append(temp_str)
    else:
        nouns.append("")
        
#as there are more than one noun in a statement
temp=[]
noun_list=[]
for i in range(0,len(nouns)):
    temp_s=nouns[i].strip()
    if nouns != "":
        temp=temp_s.split(" ")    
        noun_list.append(temp)
        temp=[]
    else:
        noun_list.append("")
    
        
#nouns tokenize
from nltk.stem.porter import PorterStemmer
ps = PorterStemmer()
temp_tokenize=[]
pos=0
temp=""
nouns_tokenize=[]
for i in range(0,len(noun_list)):
    for j in range(0,len(noun_list[i])):
        temp=ps.stem(noun_list[i][j])
        #print(temp)
        temp=temp.strip()
        if temp in bag_noun:
            print(temp)
            pos=bag_noun.index(temp)
            temp_tokenize.append(pos)
            print(pos)
        else:
            temp_tokenize.append("")
    nouns_tokenize.append(temp_tokenize)
    temp_tokenize=[]
    
    
#window_tokenize
wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\bag_of_words_1.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']

#getting all windows
s=""
window=[]
for i in range (2,48):
    s='E'+str(i)
    print(sheet_ranges[s].value)
    window.append(sheet_ranges[s].value)
    
from nltk.stem.porter import PorterStemmer
ps = PorterStemmer()
temp_tokenize=[]
pos=0
temp=""
windows_tokenize=[]
for i in range(0,len(window)):
    temp_s=window[i].strip()
    temp=temp_s.split(" ")
    print(temp)
    for k in range(0,len(temp)):     
        temp_stemmer=ps.stem(temp[k])
        print(temp_stemmer)
        if temp_stemmer in bag_of_words:
            pos=bag_of_words.index(temp_stemmer)
            temp_tokenize.append(pos)
            print(pos)
        else:
            temp_tokenize.append("")
    windows_tokenize.append(temp_tokenize)
    temp_tokenize=[]
    

#writing to excel
import xlwt
# Initialize a workbook 
book = xlwt.Workbook(encoding="utf-8")

# Add a sheet to the workbook 
sheet1 = book.add_sheet("Python Sheet 1") 




#nouns tokenized
#ADDDING 10
row = 1
col = 0
c=0
maxn=0
for i in range(0,len(nouns_tokenize)):
    for j in range(0,len(nouns_tokenize[i])):
        x=(nouns_tokenize[i][j])
        sheet1.write(row, col, x)
        col=col+1
        c=c+1
        if c>maxn:
            maxn=c
    col=0
    c=0
    row=row+1
    #row += 1
print(maxn)
    

#verbs tokenized
#ADDING 100
row = 1
col = 4
for i in range(0,len(verbs_tokenize)):
    x=verbs_tokenize[i]
    sheet1.write(row, col, x)
    row=row+1

#window tokenized
#ADDING 1000
row=1
col = 6
c=6
maxw=0
for i in range(0,len(windows_tokenize)):
    for j in range(0,len(windows_tokenize[i])):
        x=(windows_tokenize[i][j])
        sheet1.write(row, col, x)
        col=col+1
        c=c+1
        if c>maxn:
            maxw=c
    col=6
    c=6
    row=row+1
    #row += 1
print(maxw)
# Save the workbook 
book.save("spreadsheet.xls")#feature_tokens_4