#loading the dataset
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\dataset_eng.xlsx')


ws=wb.active
sheet_ranges = wb['Sheet1']

s=""
dataset=[]
for i in range (1,47):
    s='A'+str(i)
    print(sheet_ranges[s].value)
    dataset.append(sheet_ranges[s].value)
    
    
#feature extraction-unigram,nouns,window

#tokenizer for unigram
from isc_tokenizer import Tokenizer
tk = Tokenizer(lang='eng')

#tagger for tagging nouns
from isc_tagger import Tagger
tagger = Tagger(lang='eng')


#unigram
temp_u=[]
for i in range(0,len(dataset)):
    if(dataset[i] is not None):
        #print(dataset[i])
        temp_u.append(tk.tokenize(dataset[i]))
    else:
        temp_u.append("")

unigram=[]
temp=""
for i in range(0,len(temp_u)):
    for j in range(1,len(temp_u[i])):#started from 1 as we don't need indexing
        print(temp_u[i][j])
        temp=temp+temp_u[i][j]+" "
    unigram.append(temp)
    temp=""
        
                
#window
pos_of_question_word=0
posl=0
posr=0
right=[]
left=[]
right_window=[]
left_window=[]
s=0
b=0
str1="How"
str2="much"
question_word=['How','much']
str=""
window=[]
temp=""
for i in range(0,len(temp_u)):
    left=[]
    right=[]
    if temp_u[i] is not "":
        str=temp_u[i]
        #print(0)
        posl=temp_u[i].index(str1)
        posr=temp_u[i].index(str2)
        right=str[posr+1:posr+5]
        left=str[posl-3:posl]
        #print(left)
        #print(right)
        print(left+question_word+right)
        window.append(left+question_word+right)
        right_window.append(right)
        left_window.append(left)
        print(window)
        print(temp_u[i])
    else:
        window.append("")
        left_window.append("")
        right_window.append("")
        
#tagger without removing stopwords no need to run this part
noun=[]
taggers=[]
for i in range(0,len(right_window)):
    if(type(right_window[i]) is list ):
        #print(dataset[i])
        print(tagger.tag(right_window[i]))
        taggers.append(tagger.tag(right_window[i]))
    else:
        print()
        taggers.append([])
        

#removing stopwords from right window
from nltk.corpus import stopwords
#nltk.download('stopwords')
stop_words = set(stopwords.words('english'))


right_window_lower_case=[]#so that we can compare stopwords
temp=""

for j in range(0,len(right_window)):
    if right_window[j] is not None:
        #print(right_window[j])
        for i in range(0,len(right_window[j])):
            #print(right_window[j][i].lower())
            temp=temp+(right_window[j][i].lower())+" "
        right_window_lower_case.append(temp)
        temp=""
    else:
        right_window_lower_case.append()
        

right_window_without_stopwords=[]
for i in range(0,len(right_window_lower_case)):
    words=tk.tokenize(right_window_lower_case[i])
    wordsFiltered = []
    for w in words:
        if w not in stop_words:
            wordsFiltered.append(w)
    right_window_without_stopwords.append(wordsFiltered)
    

#left window without stopwords
left_window_lower_case=[]#so that we can compare stopwords
temp=""

for j in range(0,len(left_window)):
    if left_window[j] is not None:
        #print(right_window[j])
        for i in range(0,len(left_window[j])):
            #print(right_window[j][i].lower())
            temp=temp+(left_window[j][i].lower())+" "
        left_window_lower_case.append(temp)
        temp=""
    else:
        left_window_lower_case.append()
        

left_window_without_stopwords=[]
for i in range(0,len(left_window_lower_case)):
    words=tk.tokenize(left_window_lower_case[i])
    wordsFiltered = []
    for w in words:
        if w not in stop_words:
            wordsFiltered.append(w)
    left_window_without_stopwords.append(wordsFiltered)

#tagger after removing stopwords
taggers_without_stopwords=[]
for i in range(0,len(right_window_without_stopwords)):
    if(right_window_without_stopwords[i] != []):
        #print(right_window_without_stopwords[i])
        print(i)
        print(tagger.tag(right_window_without_stopwords[i]))
        taggers_without_stopwords.append(tagger.tag(right_window_without_stopwords[i]))
    else:
        #print(right_window_without_stopwords[i])
        taggers_without_stopwords.append([])
        
        
#extracting nouns and verbs
others_list=[]        
temp_n=[]
temp_v=[]
temp_o=[]
noun_list=[]
verb_list=[]
temp_str0=""
temp_str1=""
for i in range(0,len(taggers_without_stopwords)):
    for j in range(0,len(taggers_without_stopwords[i])):
        #print(i)
        print(taggers_without_stopwords[i][j])
        if(taggers_without_stopwords[i][j][1]=='NOUN'):
            #print(taggers_without_stopwords[i][j][0])
            temp_str1=taggers_without_stopwords[i][j][0]
            print(temp_str1)
            temp_n.append(taggers_without_stopwords[i][j][0]+" ")
        elif(taggers_without_stopwords[i][j][1]=='VERB'):
            temp_v.append(taggers_without_stopwords[i][j][0]+" ")
        else:
            temp_o.append(taggers_without_stopwords[i][j][0]+" ")
    noun_list.append(temp_n)
    verb_list.append(temp_v)
    others_list.append(temp_o)
    temp_n=[]
    temp_v=[]
    temp_o=[]
        

#writing to excel
import xlwt
# Initialize a workbook 
book = xlwt.Workbook(encoding="utf-8")

# Add a sheet to the workbook 
sheet1 = book.add_sheet("Python Sheet 1") 


#left window
row = 1
col = 0
temp_window=""
str_window_list_left=[]
for i in range(0,len(left_window_without_stopwords)):
    if(type(left_window_without_stopwords[i]) != [] ):
        for j in range(0,len(left_window_without_stopwords[i])):
            temp_window=temp_window+left_window_without_stopwords[i][j]+" "
        print(temp_window)
        str_window_list_left.append(temp_window)
        temp_window=""
    else:
        print(0)
        
for item in (str_window_list_left):
    sheet1.write(row, col, item) 
    row += 1
    
#right window
row = 1
col = 1
temp_window=""
str_window_list_right=[]
for i in range(0,len(right_window_without_stopwords)):
    if(type(right_window_without_stopwords[i]) != [] ):
        for j in range(0,len(right_window_without_stopwords[i])):
            temp_window=temp_window+right_window_without_stopwords[i][j]+" "
        print(temp_window)
        str_window_list_right.append(temp_window)
        temp_window=""
    else:
        print(0)
        
for item in (str_window_list_right):
    sheet1.write(row, col, item) 
    row += 1

row=1
col=2
for item in (unigram):
    sheet1.write(row, col, item) 
    row += 1    
    
row=1
col=3
for item in (noun_list):
    sheet1.write(row, col, item) 
    row += 1    

row=1
col=4
for item in (verb_list):
    sheet1.write(row, col, item) 
    row += 1    

row=1
col=5
for item in (others_list):
    sheet1.write(row, col, item) 
    row += 1    

# Save the workbook 
book.save("spreadsheet.xls")#feature_extracted_2