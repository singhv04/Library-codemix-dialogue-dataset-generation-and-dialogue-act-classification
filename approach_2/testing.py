from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\features_tokens.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']

#getting all nouns
#n1 for 1st col
#n2 for 2nd col
#ADDING 11
#if none type then its 10
#FEATURE SCALING
s=""
xn=[]
n1=[]
for i in range (2,32):
    s='A'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+11
    else:
        tx1=10
    n1.append(tx1)

n2=[]
for i in range (2,32):
    s='B'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+11
    else:
        tx1=10
    n2.append(tx1)

    
#for verbs
#if none then 100
#else +100
v1=[]
for i in range (2,32):
    s='E'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+101
    else:
        tx1=100
        print(i)
    v1.append(tx1)

#for window
#if none then 1000
#else +1001
#not +1000 as there are 0's so we won't be able to distinguish between none and 0's
w1=[]
for i in range (2,32):
    s='G'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w1.append(tx1)

w2=[]
for i in range (2,32):
    s='H'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w2.append(tx1)

w3=[]
for i in range (2,32):
    s='I'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w3.append(tx1)

w4=[]
for i in range (2,32):
    s='J'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w4.append(tx1)

w5=[]
for i in range (2,32):
    s='K'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w5.append(tx1)

w6=[]
for i in range (2,32):
    s='L'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w6.append(tx1)

w7=[]
for i in range (2,32):
    s='M'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w7.append(tx1)
    
w8=[]
for i in range (2,32):
    s='N'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w8.append(tx1)
    



#applying feature scaling
from sklearn.preprocessing import StandardScaler
sc = StandardScaler()

import numpy as np
combined_n = np.vstack((n1, n2, v1, w1, w2, w3, w4, w5, w6, w7, w8)).T
X_train = sc.fit_transform(combined_n)


#tokenizing y_train
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\dataset_1_eng.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']

#creating bag of words for test set
x=[]
test_bag=[]
for i in range (1,31):
    s='B'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x not in test_bag:
        test_bag.append(x)

#assigning temp tokens
temp_token=[]
for i in range (1,31):
    s='B'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x in test_bag:
        pos=test_bag.index(x)
        pos=pos
    temp_token.append(pos)
    
y_train=temp_token


#applying svm

from sklearn.svm import SVC
classifier = SVC(kernel = 'linear', random_state = 0)
classifier.fit(X_train, y_train)

"""
##naive bayes
from sklearn.naive_bayes import GaussianNB
classifier1 = GaussianNB()
classifier1.fit(X_train, y_train)

from sklearn.model_selection import cross_val_score
accuracies = cross_val_score(estimator = classifier1, X = X_train, y = y_train, cv = 10)
print(accuracies.mean())
print(accuracies.std())
print(accuracies)

0.645833333333
0.25259074277
[ 0.33333333  0.75        1.          0.5       ]

"""

from sklearn.model_selection import cross_val_score
accuracies = cross_val_score(estimator = classifier, X = X_train, y = y_train, cv = 4)
print(accuracies.mean())
print(accuracies.std())
print(accuracies)

"""
0.780257936508
0.130571347708
[ 0.55555556  0.875       0.85714286  0.83333333]

"""



#PREDICTION

##taking bag of words nouns and verbs
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2\\bag_of_words_3.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']


bag_of_words=[]
for i in range (2,156):
    s='A'+str(i)
    x=sheet_ranges[s].value
    bag_of_words.append(x)


bag_of_nouns=[]
for i in range (2,12):
    s='B'+str(i)
    x=sheet_ranges[s].value
    bag_of_nouns.append(x)



bag_of_verbs=[]
for i in range (2,9):
    s='C'+str(i)
    x=sheet_ranges[s].value
    bag_of_verbs.append(x)


##INPUT THE TEXT
    
test = input("enter the test")
print(test)

#tokenized
test=test.lower()
print(test)
from isc_tokenizer import Tokenizer
tk = Tokenizer(lang='hin')

tokenized=(tk.tokenize(test))
print(tokenized)

##language detection and translation

import langid  
lang_det=[]
for i in range(0,len(tokenized)):
    lang_det.append(langid.classify(tokenized[i]))
print(lang_det)

from googletrans import Translator
translator = Translator()

temp_str=""
for i in range(0,len(lang_det)):
       if lang_det[i] == 'en':
           temp_str=temp_str+tokenized[i]+" "
       else:
           translated = translator.translate(tokenized[i], dest='en')
           print("hindi word:"+tokenized[i]+" converted to english word:"+translated.text)
           temp_str=temp_str + translated.text+" "    

print(temp_str)
temp_str=temp_str.strip()



#stemming
from nltk.stem.porter import PorterStemmer
temp_list=temp_str.split(" ")
print(temp_list)

for j in range(0,len(temp_list)):
            print("before stemming:"+temp_list[j])
            ps = PorterStemmer()
            temp_list[j] = ps.stem(temp_list[j])
            print("after stemming:"+temp_list[j])
    
##tokenizing        
n_t=[]
w_t=[]
v_t=[]
l_t=[]
r_t=[]
print(temp_list)
pos=temp_list.index("how")
print(pos)
for i in range(0,len(temp_list)):
    if(i!=3 and i!=4):
        if temp_list[i] in bag_of_nouns:
            n_t.append(11+bag_of_nouns.index(temp_list[i]))
            print("noun:"+temp_list[i])
        if temp_list[i] in bag_of_verbs:
            v_t.append(101+bag_of_verbs.index(temp_list[i]))
            print("verb:"+temp_list[i])

print(temp_list)
if (len(temp_list[:pos])<3):
    start_l=len(temp_list[:pos])
    print(start_l)
else:
    start_l=3

    
left_window=temp_list[pos-start_l:pos]
print(left_window)

if (len(temp_list[pos+2:])<3):
    end_l=len(temp_list[pos+2:])
    print(end_l)
else:
    end_l=3
    

right_window=temp_list[pos+2:pos+2+end_l]
print(right_window)


for i in range(0,len(left_window)):
    if left_window[i] in bag_of_words:
        l_t.append([1001+bag_of_words.index(left_window[i])])
    else:
        l_t.append([1000])
        #print(left_window[i])
        
w_t=l_t
  
w_t.append([1001+bag_of_words.index("how")])
w_t.append([1001+bag_of_words.index("much")])

    
for i in range(0,len(right_window)):
    if right_window[i] in bag_of_words:
        #r_t.append([1000+bag_of_words.index(right_window[i])]) 
        w_t.append([1001+bag_of_words.index(right_window[i])])
    else:
        #r_t.append([1000])
        w_t.append([1000])
    

print(len(n_t))

print(len(v_t))

print(len(w_t))

n_tokens=([[i] for i in n_t])

print(n_tokens)

while len(n_tokens)<3:
    n_tokens.append([10])
    
    
v_tokens=([[i] for i in v_t])

while len(v_tokens)<1:
    v_tokens.append([100])
    
while len(w_t)<8:
    w_t.append([1000])

##prediction
combined_n_test = np.vstack((n_tokens[0], n_tokens[1], v_tokens[0], w_t[0], w_t[1], w_t[2], w_t[3], w_t[4], w_t[5], w_t[6], w_t[7])).T
X_test = sc.fit_transform(combined_n_test)
y_pred = classifier.predict(X_test)
print(y_pred)
print(test)
