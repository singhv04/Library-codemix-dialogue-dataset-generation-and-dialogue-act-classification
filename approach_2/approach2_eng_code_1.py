#the dataset along with their classification
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\dialogue act classiification\\q\\m\\kitni_m.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']
i=1
s='A'+str(i)
print(sheet_ranges[s].value)

#getting all m_kitni
s=""
m_kitni_q=[]
for i in range (1,47):
    s='A'+str(i)
    print(sheet_ranges[s].value)
    m_kitni_q.append(sheet_ranges[s].value)
    

#tokenize
from isc_tokenizer import Tokenizer
tk = Tokenizer(lang='hin')
tokenized_list=[]
for a in range(0,len(m_kitni_q)):
    temp_str=m_kitni_q[a]
    tokenized_list.append(tk.tokenize(temp_str))
    #del(tokenized_list[a][0])
   
#langdetection
import langid  
lang_det_list=[]
lang_det=[]
processed_lang_det=[]
test_str=""
for a in range(0,len(tokenized_list)):
    test_str=tokenized_list[a]
    print(test_str)
    for i in range (0,len(test_str)):
        lang_det.append(langid.classify(test_str[i]))
    processed_lang_det.append([i[0] for i in lang_det])
    lang_det=[]
        

#transliteration
#hin ="कांग्रेस party अध्यक्ष सोनिया गांधी, तमिलनाडु की मुख्यमंत्री जयललिता और रिज़र्व बैंक के गवर्नर रघुराम राजन के बीच एक समानता है. ये सभी अलग-अलग कारणों से भारतीय जनता पार्टी के राज्यसभा सांसद सुब्रमण्यम स्वामी के निशाने पर हैं. उनके जयललिता और सोनिया गांधी के पीछे पड़ने का कारण कथित भ्रष्टाचार है."
#eng = trn.transform(hin)
#print(eng)
    
#no need of transliteration we can direct translate the hindi word to english
'''transliterated=[]
for a in range(0,len(m_kitni_q)):
    temp_str=m_kitni_q[a]
    eng=trn.transform(temp_str)
    transliterated.append(eng)
'''


#translation    
from googletrans import Translator
translator = Translator()
lexical_trans=[]
temp_list=[]
temp_str=""
for a in range(0,len(processed_lang_det)):
    temp_list=processed_lang_det[a]
    #print(temp_list)
    for i in range(0,len(temp_list)):
        if temp_list[i] == 'en':
            #temp_str=temp_str+"a"
            #print(tokenized_list[a][i])
            temp_str=temp_str+tokenized_list[a][i]+" "
            #print("already english:"+tokenized_list[a][i])
        else:
            #temp_str=temp_str+"b"
            #print(tokenized_list[a][i])
            translated = translator.translate(tokenized_list[a][i], dest='en')
            print("hindi word:"+tokenized_list[a][i]+" converted to english word:"+translated.text)
            temp_str=temp_str + translated.text+" "
    lexical_trans.append(temp_str)
    temp_str=""

#manual classification
due_i=[0,1,3,4,5,32]
available_i=[2,16,18,19,20,21,22,24,28,33,34]
due_till_date_i=[6,10,40,41,42,43,44,45]
book_times_issued_i=[7,8,9,11,12]
thesis_i=[13,36,38]
book_issued_i=[15]
thesis_issued_i=[39]
total_copies_i=[17]
new_arrival_i=[23,25,26,27]
issue_limit_i=[29,30]
duration_limit_i=[31]
thesis_published_i=[35]
ambiguous_i=[14,37]

#classification manual
due=[]
for i in range(0,len(due_i)):
    due.append(lexical_trans[due_i[i]])

available=[]
for i in range(0,len(available_i)):
    available.append(lexical_trans[available_i[i]])
    
due_till_date=[]
for i in range(0,len(due_till_date_i)):
    due_till_date.append(lexical_trans[due_till_date_i[i]])
    
book_times_issued=[]
for i in range(0,len(book_times_issued_i)):
    book_times_issued.append(lexical_trans[book_times_issued_i[i]])
    
thesis=[]
for i in range(0,len(thesis_i)):
    thesis.append(lexical_trans[thesis_i[i]])

book_issued=[]
for i in range(0,len(book_issued_i)):
    book_issued.append(lexical_trans[book_issued_i[i]])
    
thesis_issued=[]
for i in range(0,len(thesis_issued_i)):
    thesis_issued.append(lexical_trans[thesis_issued_i[i]])
   
total_copies=[]
for i in range(0,len(total_copies_i)):
    total_copies.append(lexical_trans[total_copies_i[i]])

new_arrival=[]
for i in range(0,len(new_arrival_i)):
    new_arrival.append(lexical_trans[new_arrival_i[i]])

issue_limit=[]
for i in range(0,len(issue_limit_i)):
    issue_limit.append(lexical_trans[issue_limit_i[i]])

duration_limit=[]
for i in range(0,len(duration_limit_i)):
    duration_limit.append(lexical_trans[duration_limit_i[i]])

thesis_published=[]
for i in range(0,len(thesis_published_i)):
    thesis_published.append(lexical_trans[thesis_published_i[i]])

ambiguous=[]
for i in range(0,len(ambiguous_i)):
    ambiguous.append(lexical_trans[ambiguous_i[i]])


#writing to excel
import xlwt
# Initialize a workbook 
book = xlwt.Workbook(encoding="utf-8")

# Add a sheet to the workbook 
sheet1 = book.add_sheet("Python Sheet 1") 

row = 0
col = 0

for item in (due):
    sheet1.write(row, col, item)
    sheet1.write(row, col+1, 'due')
    row += 1
    

for item in (available):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'available')
    row += 1
    

for item in (due_till_date):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'due_till_date')
    row += 1


for item in (book_times_issued):
    sheet1.write(row, col, item)
    sheet1.write(row, col+1, 'book_times_issued')
    row += 1


for item in (thesis):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'thesis')
    row += 1
    
for item in (book_issued):
    sheet1.write(row, col, item)
    sheet1.write(row, col+1, 'book_issued')
    row += 1
    
for item in (thesis_issued):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'thesis_issued')
    row += 1
    
for item in (total_copies):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'total_copies')
    row += 1
      
for item in (new_arrival):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'new_arrival')
    row += 1
      
for item in (issue_limit):
    sheet1.write(row, col, item)
    sheet1.write(row, col+1, 'issue_limit')
    row += 1
        
for item in (duration_limit):
    sheet1.write(row, col, item)
    sheet1.write(row, col+1, 'duration_limit')
    row += 1
    
for item in (thesis_published):
    sheet1.write(row, col, item)
    sheet1.write(row, col+1, 'thesis_published')
    row += 1
    
for item in (ambiguous):
    sheet1.write(row, col, item) 
    sheet1.write(row, col+1, 'ambiguous')
    row += 1
    
# Save the workbook 
book.save("spreadsheet.xls")#dataset_eng