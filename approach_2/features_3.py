#reading the features
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\feature_extracted_2.xlsx')


ws=wb.active
sheet_ranges = wb['Sheet1']


#window
s=""
left_window=[]
for i in range (2,48):
    s='A'+str(i)
    print(sheet_ranges[s].value)
    left_window.append(sheet_ranges[s].value)
    
right_window=[]
for i in range (2,48):
    s='B'+str(i)
    print(sheet_ranges[s].value)
    right_window.append(sheet_ranges[s].value)
    
window=[]

for i in range(0,len(left_window)):
    temp_left=""
    temp_right=""
    temp_win=""
    if left_window[i] is not None:
        temp_left=left_window[i]
    else:
        temp_left=""
    if right_window[i] is not None:
        temp_right=right_window[i]
    else:
        temp_right=""
    temp_win=temp_left+"How much "+temp_right
    temp_win=temp_win.lower()
    temp_win=temp_win.strip()
    window.append(temp_win)
    
    
#furthrt pre-processing of window
for i in range(0,len(window)):
    print(window[i])
    window[i]=window[i].replace("mother","arrived")
    window[i]=window[i].replace("lee","taken")
    window[i]=window[i].replace("wrong","completed")
    window[i]=window[i].replace("karai","done")
    #hui,ho,


    
#provinding tokens to each noun

#importing noun
s=""
noun=[]
for i in range (2,48):
    s='D'+str(i)
    print(sheet_ranges[s].value)
    noun.append(sheet_ranges[s].value)
    
from nltk.stem.porter import PorterStemmer

#bag of words for noun
bag_noun=[]
temp=[]
for i in range(0,len(noun)):
    if noun[i] is not None:
        temp=((noun[i]).strip()).split(" ")
        print(temp)
        for j in range(0,len(temp)):
            print("before stemming:"+temp[j])
            ps = PorterStemmer()
            temp[j] = ps.stem(temp[j])
            print("after stemming:"+temp[j])
            if temp[j] not in bag_noun:
                bag_noun.append(temp[j])
                

#importing verb
s=""
verb=[]
for i in range (2,48):
    s='E'+str(i)
    print(sheet_ranges[s].value)
    verb.append(sheet_ranges[s].value)
    
#bag of words for verb
bag_verb=[]
temp=[]
for i in range(0,len(verb)):
    if verb[i] is not None:
        temp=((verb[i]).strip()).split(" ")
        #print(temp)
        for j in range(0,len(temp)):
            print("before stemming:"+temp[j])
            ps = PorterStemmer()
            temp[j] = ps.stem(temp[j])
            print("after stemming:"+temp[j])
            if temp[j] not in bag_verb:
                bag_verb.append(temp[j])
    
    
#importing bag_data
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\dataset_1_eng.xlsx')


ws=wb.active
sheet_ranges = wb['Sheet1']
    
s=""
temp=""
data=[]
for i in range (1,48):
    s='A'+str(i)
    #print(sheet_ranges[s].value)
    temp=sheet_ranges[s].value
    #print(temp)
    if temp is not None:
        pos=temp.find(" ")
        #print(pos)
        temp=temp[pos+1:]
        #print(temp)
        data.append(temp)
     
bag_data=[]
temp_x=""
temp_l=[]
for i in range(0,len(data)):
    data[i]=data[i].strip()
    temp_l=data[i].split(" ")
    #print(temp_l)
    for j in range(0,len(temp_l)):
        ps = PorterStemmer()
        temp_x=ps.stem(temp_l[j])
        #print(temp_x)
        if temp_x not in bag_data:
            bag_data.append(temp_x)


#writing to excel
import xlwt
# Initialize a workbook 
book = xlwt.Workbook(encoding="utf-8")

# Add a sheet to the workbook 
sheet1 = book.add_sheet("Python Sheet 1") 

row = 1
col = 0

for item in (bag_data):
    sheet1.write(row, col, item) 
    row += 1
    
col=1
row=1
for item in (bag_noun):
    sheet1.write(row, col, item) 
    row += 1
    
col=2
row=1    
for item in (bag_verb):
    sheet1.write(row, col, item) 
    row += 1
    
    
col=4
row=1
for item in (window):
    sheet1.write(row,col,item)
    row=row+1
# Save the workbook 
book.save("spreadsheet2.xls")#bag_of_words_3

