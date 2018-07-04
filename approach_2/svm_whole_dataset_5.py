from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\features_tokens_4.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']

#getting all nouns
#n1 for 1st col
#n2 for 2nd col
#ADDING 11
#if none type then its 10
#FEATURE SCALING
s=""
xn=[]
n1=[]
for i in range (2,48):
    s='A'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+11
    else:
        tx1=10
    n1.append(tx1)

n2=[]
for i in range (2,48):
    s='B'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+11
    else:
        tx1=10
    n2.append(tx1)

    
#for verbs
#if none then 100
#else +100
v1=[]
for i in range (2,48):
    s='E'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+101
    else:
        tx1=100
        print(i)
    v1.append(tx1)

#for window
#if none then 1000
#else +1001
#not +1000 as there are 0's so we won't be able to distinguish between none and 0's
w1=[]
for i in range (2,48):
    s='G'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w1.append(tx1)

w2=[]
for i in range (2,48):
    s='H'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w2.append(tx1)

w3=[]
for i in range (2,48):
    s='I'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w3.append(tx1)

w4=[]
for i in range (2,48):
    s='J'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w4.append(tx1)

w5=[]
for i in range (2,48):
    s='K'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w5.append(tx1)

w6=[]
for i in range (2,48):
    s='L'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w6.append(tx1)

w7=[]
for i in range (2,48):
    s='M'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w7.append(tx1)
    
w8=[]
for i in range (2,48):
    s='N'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x is not None:
        tx1=x+1001
    else:
        tx1=1000
        print(i)
    w8.append(tx1)
    



#applying feature scaling
from sklearn.preprocessing import StandardScaler
sc = StandardScaler()

import numpy as np
combined_n = np.vstack((n1, n2, v1, w1, w2, w3, w4, w5, w6, w7, w8)).T
X_train = sc.fit_transform(combined_n)


#tokenizing y_train
from openpyxl import load_workbook

wb = load_workbook('C:\\Users\\hp\\Desktop\\approach_2_mL\\dataset_1_eng.xlsx')
ws=wb.active
sheet_ranges = wb['Sheet1']

#creating bag of words for test set
x=[]
test_bag=[]
for i in range (1,47):
    s='B'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x not in test_bag:
        test_bag.append(x)

#assigning temp tokens
temp_token=[]
for i in range (1,47):
    s='B'+str(i)
    x=sheet_ranges[s].value
    #print(type(x))
    #print(x)
    if x in test_bag:
        pos=test_bag.index(x)
        pos=pos
    temp_token.append(pos)
    
y_train=temp_token


#applying svm

from sklearn.svm import SVC
classifier = SVC(kernel = 'linear', random_state = 0)
classifier.fit(X_train, y_train)


from sklearn.model_selection import cross_val_score
accuracies = cross_val_score(estimator = classifier, X = X_train, y = y_train, cv = 10)
print(accuracies.mean())
print(accuracies.std())
